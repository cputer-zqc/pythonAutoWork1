import tkinter as tk
from tkinter import filedialog
import os
from tkinter import messagebox
import pandas as pd
import openpyxl
from openpyxl.cell import _writer


class ReadAndCompareFileName:

    def __init__(self):
        self.workbook = None
        self.sheet_name = None
        self.url = None
        self.picture_number = dict()

    def get_image_path_to_excel(self, data, img_name, beforString, imgURL):
        imgURL = imgURL.replace("\\", "/") + "/"

        i = 0
        while i < len(data["名称"]):
            # 获取第i行的附件id与名称
            id_image_name = data['附件ID与名称'][i]

            # 删除掉首尾的分号
            id_image_name = id_image_name.strip(";").replace("\n", "")
            # 用分号进行分割
            id_image_name_split = id_image_name.split(";")
            # 名称拼接
            r = i + 2
            for id_image_s in id_image_name_split:
                sheet = self.workbook[self.sheet_name]
                [max_row, max_column] = self.__get_max_column_row()
                id_and_name = id_image_s.split(" ")
                image_path_value = id_and_name[1] + "(" + id_and_name[0] + ")" + ".jpg"
                filePath = imgURL + image_path_value
                # 判断最后一列是否存在数据
                if sheet.cell(row=r, column=max_column).value is None or sheet.cell(row=r,
                                                                                    column=max_column).value == "":

                    # 如果不存在数据，则将其填入其中
                    sheet.cell(row=r, column=max_column, value=os.path.relpath(filePath, self.url).replace("\\","/").replace("../", "./"))
                    self.workbook.save(self.url)
                    sheet = self.workbook[self.sheet_name]
                    sheet_data = sheet.values
                    columns = next(sheet_data)
                    data = pd.DataFrame(sheet_data, columns=columns)
                    i += 1
                else:

                    # 如果所在行的最后一列存在数据，则复制上一行的内容到新创建的行中，并修改最后一列的数据为最新的数据
                    # 获取上一行数据
                    previous_rows = list(sheet.iter_rows(min_row=r, max_row=r, values_only=True))[0]
                    # 在指定行插入与上一行相同的数据
                    sheet.insert_rows(r + 1)

                    for c_insert in range(1, sheet.max_column + 1):
                        sheet.cell(row=r + 1, column=c_insert, value=previous_rows[c_insert - 1])
                    sheet.cell(row=r + 1, column=max_column, value=os.path.relpath(filePath, self.url).replace("\\","/").replace("../", "./"))
                    self.workbook.save(self.url)
                    sheet = self.workbook[self.sheet_name]
                    sheet_data = sheet.values
                    columns = next(sheet_data)
                    data = pd.DataFrame(sheet_data, columns=columns)
                    i += 1

    def __add_emperty_column(self):
        '''
        添加空列
        :return:
        '''
        sheet = self.workbook[self.sheet_name]
        [max_row, max_column] = self.__get_max_column_row()
        # 添加空列
        for row_index in range(1, max_row + 1):
            if row_index == 1:
                sheet.cell(row=row_index, column=max_column + 1, value='point_img_id')  # 稍后修改
            else:
                sheet.cell(row=row_index, column=max_column + 1, value=None)
        self.workbook.save(self.url)

    def add_point_image_id(self, point_img_id_name, point_img_id_value):

        self.__add_emperty_column()
        [max_row, max_column] = self.__get_max_column_row()
        number = 1  # 稍后修改
        sheet = self.workbook[self.sheet_name]
        sheet_data = sheet.values
        columns = next(sheet_data)
        data = pd.DataFrame(sheet_data, columns=columns)
        i = 0
        while i < len(data['附件ID与名称']):
            r = i + 2
            if i == 0:
                sheet.cell(row=r, column=max_column, value=number)
                self.workbook.save(self.url)
                sheet_data = sheet.values
                columns = next(sheet_data)
                data = pd.DataFrame(sheet_data, columns=columns)
                i += 1
                continue
            if data['附件ID与名称'][i] == data['附件ID与名称'][i - 1]:
                sheet.cell(row=r, column=max_column, value=number)
                self.workbook.save(self.url)
                sheet_data = sheet.values
                columns = next(sheet_data)
                data = pd.DataFrame(sheet_data, columns=columns)
                i += 1
            else:
                number += 1
                sheet.cell(row=r, column=max_column, value=number)
                self.workbook.save(self.url)
                sheet_data = sheet.values
                columns = next(sheet_data)
                data = pd.DataFrame(sheet_data, columns=columns)
                i += 1

    def __get_max_column_row(self):
        '''
        获取当前表格醉倒列与最大行数
        :return: {"mr":max_row, "mc":max_column}
        '''
        sheet = self.workbook[self.sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        return [max_row, max_column]

    def get_data(self, url, addName):
        '''
        获取表格中信息
        :param url:
        :return:
        '''
        self.url = url
        # 获取表格
        self.workbook = openpyxl.load_workbook(self.url)

        # 获取sheet
        self.sheet_name = self.workbook.sheetnames[0]
        sheet = self.workbook[self.sheet_name]
        row_data = []
        for cell in sheet[1]:
            row_data.append(cell.value)
        if addName not in row_data[1]:
            [max_row, max_column] = self.__get_max_column_row()
            # 添加空列
            for row_index in range(1, max_row + 1):
                if row_index == 1:
                    sheet.cell(row=row_index, column=max_column + 1, value=addName)
                else:
                    sheet.cell(row=row_index, column=max_column + 1, value=None)
        self.workbook.save(url)
        # 数据转换
        data = sheet.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        # 返回数据
        return df

    def my_close_workbook(self, value):
        # sheet = self.workbook[self.sheet_name]
        # if value in sheet[1]:
        #     [index_row, index_column] = self.__get_column_row_number(value)
        #     sheet.delete_cols(index_column)
        self.workbook.close()


# 创建主窗口
root = tk.Tk()
root.title("文件加载与转换程序")


# 定义函数：打开目录选择对话框，并在输入框中显示文件夹路径
def open_folder_dialog(entry_var):
    folder_path = filedialog.askdirectory()  # 打开目录选择对话框
    if folder_path:  # 如果用户选择了目录
        entry_var.set(folder_path)  # 在输入框中显示目录路径


# 定义函数：打开文件选择对话框，并在输入框中显示文件路径
def open_file_dialog(entry_var):
    file_path = filedialog.askopenfilename()  # 打开文件选择对话框
    if file_path:  # 如果用户选择了文件
        entry_var.set(file_path)  # 在输入框中显示文件路径


# 定义函数：开始转换的操作（示例函数，需要根据实际需求编写）
def start_conversion():
    entry_var2.set("E:\项目文件夹\软件开发类\江宁市政设施调查平台开发报价\江宁普查项目外业资料\测试资料/10.16调研.csv")
    entry_var1.set("E:\项目文件夹\软件开发类\江宁市政设施调查平台开发报价\江宁普查项目外业资料\测试资料/10.16调研")
    if entry_var2.get() == None or entry_var2.get() == "":
        messagebox.showinfo("woring！！！", "请选择.xls、.xlsx、.csv文件")
    elif entry_var1.get() == None or entry_var1.get() == "":
        messagebox.showinfo("woring！！！", "请选择照片所在文件夹")
    else:
        # 进行文件转换操作
        url_excel = None
        if entry_var2.get().endswith(".csv"):
            # 读取CSV文件
            csv_file = entry_var2.get()
            df = pd.read_csv(csv_file, encoding='gbk')
            # 将DataFrame保存为Excel文件
            url_excel = entry_var2.get().replace(".csv", "") + '.xlsx'
            df.to_excel(url_excel, index=False)
        elif entry_var2.get().endswith(".xlsx"):
            url_excel = entry_var2.get()
        elif entry_var2.get().endswith(".xls"):
            url_excel = entry_var2.get()
        else:
            messagebox.showinfo("woring！！！", "无法打开所选文件，请重新选择")
        rcf = ReadAndCompareFileName()
        addName = add_column_label_entry.get()
        if addName == "" or addName is None:
            addName = "img_path"
        data = rcf.get_data(url_excel, addName)
        # 获取目录下文件名称
        url_img = entry_var1.get()
        img_name = os.listdir(url_img)

        beforString = entry_var3.get()
        if beforString == "" or beforString is None:
            beforString = "pic_"
        else:
            beforString += "_"
        rcf.get_image_path_to_excel(data, img_name, beforString, url_img)
        rcf.add_point_image_id('point_img_id_name', 'point_img_id_value')
        rcf.my_close_workbook("excelTimeCode")
        messagebox.showinfo("转换完成", url_excel + "转换已完成！")
def split_road_by_subject():
    '''
    按照专业对道路进行分割,
    同一个断面只有一个分割线。
    :return:
    '''

    
    pass

# Excel文件路径
# 创建标签、输入框和按钮（图片目录）
label1 = tk.Label(root, text="图片目录路径:")
label1.grid(row=0, column=0, padx=10, pady=10, sticky="w")  # 放置在第0行第0列

entry_var1 = tk.StringVar()
entry1 = tk.Entry(root, textvariable=entry_var1, width=50)
entry1.grid(row=0, column=1, padx=10, pady=10, sticky="w")  # 放置在第0行第1列

button1 = tk.Button(root, text="①选择图片目录", command=lambda: open_folder_dialog(entry_var1))
button1.grid(row=0, column=2, padx=10, pady=10, sticky="w")  # 放置在第0行第2列

# image存放的路径
label2 = tk.Label(root, text="文件路径:")
label2.grid(row=1, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

entry_var2 = tk.StringVar()
entry2 = tk.Entry(root, textvariable=entry_var2, width=50)
entry2.grid(row=1, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

button2 = tk.Button(root, text="②选择文件", command=lambda: open_file_dialog(entry_var2))
button2.grid(row=1, column=2, padx=10, pady=10, sticky="w")  # 放置在第1行第2列


# 分割线文件路径
splitLineFilePath = tk.Label(root, text="分割线文件路径:")
splitLineFilePath.grid(row=2, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

split_file_var = tk.StringVar()
split_file_entery = tk.Entry(root, textvariable=split_file_var, width=50)
split_file_entery.grid(row=2, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

split_file_button = tk.Button(root, text="②选择分割线文件", command=lambda: open_file_dialog(split_file_entery))
split_file_button.grid(row=2, column=2, padx=10, pady=10, sticky="w")  # 放置在第1行第2列

# image存放的路径
label3 = tk.Label(root, text="图片名前缀:")
label3.grid(row=3, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

entry_var3 = tk.StringVar()
entry3 = tk.Entry(root, textvariable=entry_var3, width=50)
entry3.grid(row=3, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

# 添加列的名称
add_column_label = tk.Label(root, text="表格中图片所在列名称:")
add_column_label.grid(row=4, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

add_column_label_entry_var = tk.StringVar()
add_column_label_entry = tk.Entry(root, textvariable=add_column_label_entry_var, width=50)
add_column_label_entry.grid(row=4, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

# 最下方按钮
# 创建开始转换按钮和关闭程序按钮（位于同一行）
start_button = tk.Button(root, text="③开始转换", command=start_conversion)
start_button.grid(row=5, column=0, padx=10, pady=10, sticky="w")  # 放置在第2行第0列

# 最下方按钮
# 创建开始转换按钮和关闭程序按钮（位于同一行）
start_button = tk.Button(root, text="④路段切分", command=split_road_by_subject)
start_button.grid(row=5, column=1, padx=10, pady=10, sticky="w")  # 放置在第2行第0列

close_button = tk.Button(root, text="关闭程序", command=root.quit)
close_button.grid(row=5, column=2, padx=10, pady=10, sticky="w")  # 放置在第2行第1列
# # 加载Excel表格按钮
# load_excel_button = tk.Button(root, text="Load Excel", command=load_excel)
# load_excel_button.pack(side=tk.LEFT)
# 运行主循环
root.mainloop()
