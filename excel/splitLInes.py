import openpyxl
import pandas as pd
import os


class split_lines:

    def __init__(self):
        self.workbook = None
        self.sheet_name = None
        self.url = None
        self.picture_number = dict()

    def __get_max_column_row(self):
        '''
        获取当前表格最大列与最大行数
        :return: {"mr":max_row, "mc":max_column}
        '''
        sheet = self.workbook[self.sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        return [max_row, max_column]

    def __get_column_row_value(self, value):
        '''
        查找具有指定元素的单元格
        :param value:
        :return:
        '''
        sheet = self.workbook[self.sheet_name]
        # 指定要查找的元素
        target_element = value
        # 遍历工作表，查找具有指定元素的单元格
        for row_index, row in enumerate(sheet.iter_rows(), start=1):
            for col_index, cell in enumerate(row, start=1):
                if cell.value == target_element:
                    # 找到了具有指定元素的单元格
                    return [row_index, col_index]

    def __csv_to_xlsx(self, url):
        if url.endswith('.csv'):

            csv_url = url
            df = pd.read_csv(csv_url, encoding='gbk')
            xlsx_url = url.replace('.csv', '.xlsx')
            df.to_excel(xlsx_url, index=False)
            # 将DataFrame保存为Excel文件
            return xlsx_url
        elif url.endswith('.xls'):
            return url
        elif url.endswith('.xlsx'):
            return url
        else:
            print('无法打开请重新选择')

    def get_data(self, url):
        '''
        获取表格中信息
        :param url:
        :return:
        '''

        self.url = self.__csv_to_xlsx(url)
        if self.url is None: return '请选择正确的文件'

        # 获取表格
        self.workbook = openpyxl.load_workbook(self.url)

        # 获取sheet
        self.sheet_name = self.workbook.sheetnames[0]
        sheet = self.workbook[self.sheet_name]

        # 数据转换
        data = sheet.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        # 返回数据
        return df

    def get_row_subject(self, df_data):
        '''
        获取一张表格中不同的专业的分割线
            1、获取所有道路的名称
        :return:
        '''
        road_name = dict()
        subject_name = dict()
        for i in range(len(df_data)):
            split_divide_line = df_data['名称'][i].split("-")
            # 获取表格中所有路名
            if len(road_name) == 0:
                road_name[i] = split_divide_line[0]
            else:
                if split_divide_line[0] not in road_name.values():
                    road_name[i] = split_divide_line[0]
            # 获取表格中所有专业名
            if len(subject_name) == 0:
                subject_name[i] = split_divide_line[0]
            else:
                if split_divide_line[0] not in subject_name.values():
                    subject_name[i] = split_divide_line[0]

        # 获取xlsx文件名称
        filenamelist = os.listdir('./doc')
        for i in range(len(filenamelist)):
            
            pass
        pass

    def my_close_workbook(self, value):
        '''
        关闭表格
        :param value:
        :return:
        '''
        sheet = self.workbook[self.sheet_name]
        if value in sheet[1]:
            [index_row, index_column] = self.__get_column_row_number(value)
            sheet.delete_cols(index_column)
        self.workbook.close()


if __name__ == "__main__":

    '''
    1、获取分割线文件所在路径
    2、打开文件
    3、匹配专业
    4、计算方向
    5、存入xlsx表格
    '''

    # 获取Excel表格中数据
    url_excel = 'E:/01code/pythonAutoWork/excel/doc/普查-导出文件.csv'
    sl = split_lines()
    table_data = sl.get_data(url_excel)
    print(table_data)