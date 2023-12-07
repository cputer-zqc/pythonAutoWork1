import shutil
import tkinter as tk
from tkinter import filedialog
import os
from tkinter import messagebox
from tkinter import ttk

import numpy as np
import pandas as pd
import openpyxl
import chardet
import re
from openpyxl.styles import PatternFill

from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from openpyxl.cell import _writer

'''
天地图：dUfD7ZHeYbe3
'''


class ReadAndCompareFileName:

    def __init__(self):
        self.workbook = None
        self.sheet_name = None
        self.url = None
        self.picture_number = dict()
        self.input_field = {
                            "信号灯": "机动车信号灯",
                            "信号": "机动车信号灯",
                            "交通信号灯": "机动车信号灯",
                            "交通信号灯立柱": "机动车信号灯",

                            "人行灯": '人行信号灯',
                            "行人灯": '人行信号灯',
                            "人行信号灯": '人行信号灯',
                            "入行灯": '人行信号灯',
                            "人形灯": '人行信号灯',

                            "指路": "指路标志",
                            "标志牌：指路牌": "指路标志",
                            "指路牌": "指路标志",

                            "分道": "车道行驶标志",
                            "标志牌：分道牌": "车道行驶标志",
                            "分道牌": "车道行驶标志",
                            "车道指示标志": "车道行驶标志",
                            "车道": "车道行驶标志",

                            "路名": "路名牌",
                            "标志牌：路名牌": "路名牌",
                            "跟名牌": "路名牌",
                            "跟铭牌": "路名牌",


                            "桥名": "桥名牌",

                            "禁令": "禁令标志",
                            "标志牌：禁令": "禁令标志",
                            "禁令牌": "禁令标志",
                            "组合禁令": "禁令标志",
                            "禁令牌及其它": "禁令标志",
                            "禁停标志牌": "禁令标志",
                            "组合禁令，含单块禁停": "禁令标志",

                            "机非": "机非分道行驶标志",
                            "标志牌：机非": "机非分道行驶标志",

                            "限高": "限高标志",
                            "限高牌": "限高标志",
                            "标志牌：限高": "限高标志",
                            "限载": "限载标志",
                            "标志牌：限重": "限载标志",
                            "限载牌": "限载标志",

                            "门架": "限高架",

                            "清拖": "禁停清拖标志",
                            "标志牌：清拖": "禁停清拖标志",
                            "禁停清拖牌": "禁停清拖标志",
                            "违停清拖": "禁停清拖标志",

                            "其他": "其他道路设施",
                            "标志牌：其他抓拍时间": "其他道路设施",
                            "标志牌：其他文明标语": "其他道路设施",
                            "标志牌：文明标语": "其他道路设施",
                            "其他：文明标语": "其他道路设施",
                            "标志牌：其他": "其他道路设施",
                            "标志牌：其他潮汐车道": "其他道路设施",
                            "其它": "其他道路设施",
                            "人行过街": "其他道路设施",
                            "标志牌：其他应急避难场所": "其他道路设施",
                            "标志牌其他": "其他道路设施",
                            }

    def get_image_path_to_excel(self, data, img_name, imgURL):
        '''
        向excel表格中填充img的路径
        :param data:
        :param img_name:
        :param imgURL:
        :return:
        '''
        imgURL = imgURL.replace("\\", "/") + "/"
        i = 0
        while i < len(data["附件ID与名称"]):

            # 获取第i行的附件id与名称
            id_image_name = data['附件ID与名称'][i]
            # 删除掉首尾的分号
            id_image_name = id_image_name.strip(";").replace("\n", "")
            # 用分号进行分割
            id_image_name_split = id_image_name.split(";")
            # 名称拼接
            r = i + 2
            for id_image_s in id_image_name_split:
                sheet = self.workbook[self.sheet_name]
                [max_row, max_column] = self.__get_max_column_row()
                id_and_name = id_image_s.split(" ")
                image_path_value = "?????"
                if len(id_and_name) >= 2:
                    image_path_value = id_and_name[1] + "(" + id_and_name[0] + ")" + ".jpg"
                else:
                    print(id_image_s)
                filePath = imgURL + image_path_value
                # 判断最后一列是否存在数据
                img_path = os.path.relpath(filePath, self.url).replace("\\", "/").replace("../", "./")
                if sheet.cell(row=r, column=max_column).value is None or sheet.cell(row=r,
                                                                                    column=max_column).value == "":
                    # 如果不存在数据，则将其填入其中
                    sheet.cell(row=r, column=max_column,
                               value=img_path)
                    sheet.cell(row=r, column=max_column).hyperlink = f'file:///{filePath}'
                    # self.workbook.save(self.url)
                    sheet = self.workbook[self.sheet_name]
                    sheet_data = sheet.values
                    columns = next(sheet_data)
                    data = pd.DataFrame(sheet_data, columns=columns)
                    i += 1
                else:
                    # 如果所在行的最后一列存在数据，则复制上一行的内容到新创建的行中，并修改最后一列的数据为最新的数据
                    # 获取上一行数据
                    previous_rows = list(sheet.iter_rows(min_row=r, max_row=r, values_only=True))[0]
                    # 在指定行插入与上一行相同的数据
                    sheet.insert_rows(r + 1)
                    for c_insert in range(1, sheet.max_column + 1):
                        sheet.cell(row=r + 1, column=c_insert, value=previous_rows[c_insert - 1])
                    sheet.cell(row=r + 1, column=max_column,
                               value=img_path)
                    sheet.cell(row=r + 1, column=max_column).hyperlink = f'file:///{filePath}'
                    # self.workbook.save(self.url)
                    sheet = self.workbook[self.sheet_name]
                    sheet_data = sheet.values
                    columns = next(sheet_data)
                    data = pd.DataFrame(sheet_data, columns=columns)
                    r += 1
                    i += 1
        self.workbook.save(self.url)

    def __add_emperty_column(self):
        '''
        添加空列
        :return:
        '''
        sheet = self.workbook[self.sheet_name]
        [max_row, max_column] = self.__get_max_column_row()
        # 添加空列
        for row_index in range(1, max_row + 1):
            if row_index == 1:
                sheet.cell(row=row_index, column=max_column + 1, value='point_img_id')  # 稍后修改
            else:
                sheet.cell(row=row_index, column=max_column + 1, value=None)
        self.workbook.save(self.url)

    def detect_encoding(self, file_path):
        with open(file_path, 'rb') as f:
            detector = chardet.universaldetector.UniversalDetector()
            for line in f.readlines():
                detector.feed(line)
                if detector.done:
                    break
            detector.close()
            encoding = detector.result['encoding']
        return encoding

    def add_point_image_id(self, point_img_id_name, point_img_id_value):
        '''
        获取
        :param point_img_id_name:
        :param point_img_id_value:
        :return:
        '''
        self.__add_emperty_column()
        [max_row, max_column] = self.__get_max_column_row()
        number = 1  # 稍后修改
        sheet = self.workbook[self.sheet_name]
        sheet_data = sheet.values
        columns = next(sheet_data)
        data = pd.DataFrame(sheet_data, columns=columns)
        i = 0
        while i < len(data['附件ID与名称']):
            r = i + 2
            if i == 0:
                sheet.cell(row=r, column=max_column, value=number)
                # self.workbook.save(self.url)
                sheet_data = sheet.values
                columns = next(sheet_data)
                data = pd.DataFrame(sheet_data, columns=columns)
                i += 1
                continue
            if data['附件ID与名称'][i] == data['附件ID与名称'][i - 1]:
                sheet.cell(row=r, column=max_column, value=number)
                # self.workbook.save(self.url)
                sheet_data = sheet.values
                columns = next(sheet_data)
                data = pd.DataFrame(sheet_data, columns=columns)
                i += 1
            else:
                number += 1
                sheet.cell(row=r, column=max_column, value=number)
                # self.workbook.save(self.url)
                sheet_data = sheet.values
                columns = next(sheet_data)
                data = pd.DataFrame(sheet_data, columns=columns)
                i += 1
        self.workbook.save(self.url)

    def field_matching(self):
        '''
        自动匹配，市政设施备注与正式名称
        :return:
        '''
        sheet = self.workbook[self.sheet_name]
        # 数据转换
        data = sheet.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        column_number = -1
        for c in range(1, sheet.max_column + 1):
            if 'Comment' == sheet.cell(row=1, column=c).value:
                column_number = c
        for i in range(len(df['Comment'])):
            if df['Comment'][i] in self.input_field.keys():
                sheet.cell(row=i + 2, column=column_number, value=self.input_field[df['Comment'][i]])
        self.workbook.save(self.url)

    def __get_comment_column_number(self):
        sheet = self.workbook[self.sheet_name]
        # 添加三列
        # 数据转换
        data = sheet.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        column_number = -1
        for c in range(1, sheet.max_column):
            if 'Comment' == sheet.cell(row=1, column=c).value:
                column_number = c
        return column_number

    def add_ludeng_column(self):
        '''
        向路灯表格中添加列信息
        :return:
        '''
        sheet = self.workbook[self.sheet_name]
        # 添加三列
        # 数据转换
        column_number = self.__get_comment_column_number()
        # 插入三列
        # 添加单侧双侧
        # 添加标牌颜色
        sheet.insert_cols(column_number, amount=3)
        column_name = {"0": "布置方式", "1": "路灯铭牌（绿色/蓝色/其他）", "2": "路灯盏数（盏/根）"}
        for i in range(3):
            sheet.cell(row=1, column=column_number + i, value=column_name[str(i)])
        self.workbook.save(self.url)
        sheet = self.workbook[self.sheet_name]

        # 添加灯头数
        column_number = self.__get_comment_column_number()
        sheet = self.workbook[self.sheet_name]
        for r in range(2, sheet.max_row + 1):
            # 获取路灯数
            value = sheet.cell(row=r, column=column_number).value
            numbers = list()
            if value is not None and value != "":
                if  isinstance(value, int):
                    numbers.append(value)
                else:
                    numbers = re.findall(r'\d', value)
            else:
                numbers.append(str(-1))
            if len(numbers) == 0:
                sheet.cell(row=r, column=column_number - 1, value=str(-1))
            else:
                if sheet.cell(row=r, column=column_number).value is not None:
                    value = None
                    if isinstance(sheet.cell(row=r, column=column_number).value, int):
                        value = None
                    else:
                        value = ''.join(char for char in sheet.cell(row=r, column=column_number).value if char.isalpha())
                    sheet.cell(row=r, column=column_number, value=value)
                sheet.cell(row=r, column=column_number - 1, value=numbers[0])
        self.workbook.save(self.url)

        # 添加在线情况，max_column-2
        sheet = self.workbook[self.sheet_name]
        column_number = self.__get_comment_column_number()
        sheet.insert_cols(column_number + 1)
        for r in range(2, sheet.max_row + 1):
            if sheet.cell(row=r, column=column_number).value is not None and sheet.cell(row=r, column=column_number).value != "":
                if isinstance(sheet.cell(row=r, column=column_number).value, int):
                    sheet.cell(row=r, column=column_number, value="功能灯")
                    sheet.cell(row=r, column=column_number + 1, value="在线")
                else:
                    if "损坏" not in sheet.cell(row=r, column=column_number).value:
                        sheet.cell(row=r, column=column_number + 1, value="在线")
                    else:
                        value = sheet.cell(row=r, column=column_number).value
                        if value is not None and value != "":
                            value = value.replace("损坏", "")
                            if "," in value:
                                value = value.replace(",", "")
                            elif "，":
                                value = value.replace("，", "")
                        if value == "":
                            sheet.cell(row=r, column=column_number, value="功能灯")
                        else:
                            sheet.cell(row=r, column=column_number, value=value)
                        sheet.cell(row=r, column=column_number + 1, value="损坏")
            else:
                sheet.cell(row=r, column=column_number, value="功能灯")
                sheet.cell(row=r, column=column_number + 1, value="在线")

        sheet.cell(row=1, column=column_number + 1, value="在线状态")

        self.workbook.save(self.url)

    def set_back_color(self):
        # 创建一个填充样式，设置背景颜色为黄色
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet = self.workbook[self.sheet_name]
        column_number = self.__get_comment_column_number()
        for r in range(2, sheet.max_row + 1):
            if sheet.cell(row=r, column=column_number - 1).value == "-1":
                sheet.cell(row=r, column=column_number - 1).fill = yellow_fill
        self.workbook.save(self.url)

    def add_belong_to(self, belong):
        '''
        添加管养区域
        :param belong:
        :return:
        '''
        sheet = self.workbook[self.sheet_name]
        sheet.insert_cols(2)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.value = belong
        for i in range(2, sheet.max_row + 1):
            value = sheet.cell(row=i, column=1).value.strip("/")
            if "环卫-" in value:
                value = value.replace("环卫-", "")
            if "环卫 " in value:
                value = value.replace("环卫 ", "")
            if "（补检查井）" in value:
                value = value.replace("（补检查井）", "")
            if "环卫—" in value:
                value = value.replace("环卫—", "")
            if "环卫" in value:
                value = value.replace("环卫", "")
            if "文华街（合并）/" in value:
                value = value.replace("文华街（合并）/","")
            if "、" in value:
                value = value.replace("、","-")
            if "（少一段）" in value:
                value = value.replace("（少一段）", "")
            if "（像单边）" in value:
                value = value.replace("（像单边）","")
            if "319" in value:
                value = value.replace("319","")
            if "816（多一段滨河东路）" in value:
                value = value.replace("816（多一段滨河东路）", "")
            if "（井太少）" in value:
                value = value.replace("（井太少）","")
            if "照明-" in value:
                value = value.replace("照明-", "")
            if "照明 " in value:
                value = value.replace("照明 ","")
            if "照明" in value:
                value= value.replace("照明","")
            if "补点" in value:
                value= value.replace("补点","")
            if "核定-" in value:
                value = value.replace("核定-", "")
            if "绿化-" in value:
                value = value.replace("绿化-","")
            if "绿化  " in value:
                value = value.replace("绿化  ", "")
            if "绿化数据-" in value:
                value = value.replace("绿化数据-", "")
            if "交通-" in value:
                value = value.replace("交通-", "")
            sheet.cell(row=i, column=1, value=value)
        sheet.cell(row=1, column=2, value="管养区域")
        sheet.cell(row=1, column=1, value="道路名称")
        self.workbook.save(self.url)

    def __get_max_column_row(self):
        '''
        获取当前表格醉倒列与最大行数
        :return: {"mr":max_row, "mc":max_column}
        '''
        sheet = self.workbook[self.sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        return [max_row, max_column]

    def get_data(self, url, addName):
        '''
        获取表格中信息
        :param url:
        :return:
        '''
        self.url = url
        # 获取表格
        self.workbook = openpyxl.load_workbook(self.url)

        # 获取sheet
        self.sheet_name = self.workbook.sheetnames[0]
        sheet = self.workbook[self.sheet_name]
        row_data = []
        for cell in sheet[1]:
            row_data.append(cell.value)
        if addName not in row_data[1]:
            [max_row, max_column] = self.__get_max_column_row()
            # 添加空列
            for row_index in range(1, max_row + 1):
                if row_index == 1:
                    sheet.cell(row=row_index, column=max_column + 1, value=addName)
                else:
                    sheet.cell(row=row_index, column=max_column + 1, value=None)
        self.workbook.save(url)
        # 数据转换
        data = sheet.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        # 返回数据
        return df

    def read_excel(self, excel_file):
        '''
        获取excel表格中道路线路坐标并将其保存到csv文件中
        :param excel_file:
        :return:
        '''
        workbook = openpyxl.load_workbook(excel_file)
        # 获取sheet
        sheet_name = workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        data = dict()
        # {文件夹名称}
        column = ("文件夹", "名称", "经纬度[经度 + 纬度]", "线条宽度", "0XFA0000FF", "线条不透明度", "闭合", "线型", "轨迹风格", "Comment")
        data["文件夹"] = list()
        data["名称"] = list()
        data["经纬度[经度 + 纬度]"] = list()
        data["线条宽度"] = list()
        data["线条颜色"] = list()
        data["线条不透明度"] = list()
        data["闭合"] = list()
        data["线型"] = list()
        data["轨迹风格"] = list()
        data["Comment"] = list()
        for r in range(5, sheet.max_row +1):
            # 列的内容为指定列的内容
            # 2、sheet.max_row-1
            data["文件夹"].append("")
            road_name = sheet.cell(row=r, column=2).value
            data["名称"].append(road_name)
            coordinate = sheet.cell(row=r, column=sheet.max_column-1).value
            line = str()
            if coordinate is not None and coordinate != "":
                # if "|" in coordinate:
                coordinate = coordinate.replace("POINT|","")
                coordinate = coordinate.replace("OINT|","")
                coordinate = coordinate.replace("INT|","")
                coordinate = coordinate.replace("NT|","")
                coordinate = coordinate.replace("T|","")
                coordinate = coordinate.replace("POLYLINE|","")
                coordinate = coordinate.replace("OLYLINE|","")
                coordinate = coordinate.replace("LYLINE|","")
                coordinate = coordinate.replace("YLINE|","")
                coordinate = coordinate.replace("LINE|","")
                coordinate = coordinate.replace("INE|","")
                coordinate = coordinate.replace("NE|","")
                coordinate = coordinate.replace("E|","")
                coordinate = coordinate.replace("|","")
                position = coordinate.split(";")
                for p in range(len(position)):
                    pos = position[p].split(",")
                    x = pos[0]
                    y = pos[1]
                    po = str(y) +"," + str(x) + ";"
                    line += po
                # else:
                #     pass

            data["经纬度[经度 + 纬度]"].append(line.strip(";"))
            data["线条宽度"].append("8")
            if line.strip(";") == "" or line.strip(";") is None or line.strip(";") == 'nan':
                data["线条颜色"].append("0XFA53FFEE")
                poline = data["经纬度[经度 + 纬度]"][len(data["经纬度[经度 + 纬度]"]) - 2]
                poline_points = poline.split(";")

                for point in range(len(poline_points)):
                    station = poline_points[point].split(",")
                    staion_x = float(station[0]) + 0.0005
                    staion_y = float(station[1]) + 0.0005
                    line += str(staion_x) + "," + str(staion_y) + ";"
                    pass
                data["经纬度[经度 + 纬度]"][len(data["经纬度[经度 + 纬度]"])-1] = line.strip(";")
            else:
                data["线条颜色"].append("0XFA0000FF")
            data["线条不透明度"].append("50")
            data["闭合"].append("0")
            data["线型"].append("0")
            data["轨迹风格"].append("4")
            data["Comment"].append(sheet.cell(row=r, column= 1).value)

        df = pd.DataFrame(data)

        current_directory = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_directory, excel_file)
        parent_directory = os.path.dirname(file_path)
        (file, ext) = os.path.splitext(excel_file)
        (path, fileName) = os.path.split(excel_file)
        fileName = fileName.replace(ext,"")
        df.to_csv(parent_directory + "/" + fileName +".csv",index =False, encoding='gbk')

    def read_excel2(self,excel_file):
        '''
        将坐标转换为平台坐标
        :param excel_file:
        :return:
        '''
        pd_data_old = pd.read_excel(excel_file)
        pd_data_new  = pd_data_old.copy()
        mid_point = dict()
        mid_poit_list = list()
        for i in range(len(pd_data_old["经纬度[经度 + 纬度]"])):
            old_data = pd_data_old["经纬度[经度 + 纬度]"][i]
            coordation_new = str()
            if type(old_data) is not str:
                old_data = str(old_data)
            if old_data is not None and old_data != "" and old_data != 'nan' and old_data != 'None':
                old_data_split_semicolon = old_data.split(";")
                for j in range(len(old_data_split_semicolon)):
                    coordation_old = old_data_split_semicolon[j].split(",")
                    lon = coordation_old[0]
                    lat = coordation_old[1]
                    if j == 0:
                        coordation_new = "POLYLINE|" + lat + "," + lon + ";"
                    elif j == len(old_data_split_semicolon) -1:
                        coordation_new += lat + "," + lon
                    else:
                        coordation_new += lat + "," + lon + ";"
                len_number = len(old_data_split_semicolon) % 2
                midpoint = str()
                if len_number == 1:
                    midnumber = np.ceil(len_number/2)
                    point = old_data_split_semicolon[int(midnumber)].split(",")
                    midpoint = "point|" + point[1] + "," + point[0]
                elif len_number == 0:
                    midnumber = np.ceil(len_number / 2)
                    point1 = old_data_split_semicolon[int(midnumber)].split(",")
                    point2 = old_data_split_semicolon[int(midnumber) + 1].split(",")
                    midpoint_lon = (float(point1[1]) + float(point2[1])) / 2
                    midpoint_lat = (float(point1[0]) + float(point2[0])) / 2
                    midpoint = "point|" + str(midpoint_lon) + "," + str(midpoint_lat)
                mid_poit_list.append(midpoint)
            else:
                coordation_new = None
                mid_poit_list.append(None)
            pd_data_new.loc[i,"经纬度[经度 + 纬度]"] = coordation_new
        pd_data_new['中心点'] = mid_poit_list
        # other_excel = ""
        # workbook = openpyxl.load_workbook(other_excel)
        # sheetname = workbook[0]
        #
        # self.__get_comment_column_number()
        pd_data_new.to_excel(excel_file,index =False)

    def my_close_workbook(self, value):
        self.workbook.close()

    def clean_huanwei_column(self):
        '''
        格式化绿化Comment
        :return:
        '''
        sheet = self.workbook[self.sheet_name]
        commnet_column_number = self.__get_comment_column_number()
        for r in range(2, sheet.max_row+1):
            if sheet.cell(row=r, column= commnet_column_number).value is not None and sheet.cell(row=r, column= commnet_column_number).value != "":
                if "井" in sheet.cell(row=r, column= commnet_column_number).value:
                    sheet.cell(row=r, column=commnet_column_number,value="检查井")
                elif "井盖" in sheet.cell(row=r, column= commnet_column_number).value:
                    sheet.cell(row=r, column=commnet_column_number, value="检查井")
                elif "景" in sheet.cell(row=r, column= commnet_column_number).value:
                    sheet.cell(row=r, column=commnet_column_number, value="检查井")
                elif "雨" in sheet.cell(row=r, column= commnet_column_number).value:
                    sheet.cell(row=r, column=commnet_column_number, value="检查井")
                elif "垃圾" in sheet.cell(row=r, column= commnet_column_number).value:
                    sheet.cell(row=r, column=commnet_column_number, value="垃圾箱")
                elif "篦子" in sheet.cell(row=r, column= commnet_column_number).value:
                    sheet.cell(row=r, column=commnet_column_number, value="雨水井")
                elif "蓖子" in sheet.cell(row=r, column= commnet_column_number).value:
                    sheet.cell(row=r, column=commnet_column_number, value="雨水井")

                elif "鼻子" in sheet.cell(row=r, column= commnet_column_number).value:
                    sheet.cell(row=r, column=commnet_column_number, value="雨水井")
                elif "作业" in sheet.cell(row=r, column= commnet_column_number).value:
                    sheet.cell(row=r, column=commnet_column_number, value="座椅")
                elif "公厕 无公示" in sheet.cell(row=r, column= commnet_column_number).value:
                    sheet.cell(row=r, column=commnet_column_number, value="公厕")
                elif "厕所" in sheet.cell(row=r, column= commnet_column_number).value:
                    sheet.cell(row=r, column=commnet_column_number, value="公厕")
            else:
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                sheet.cell(row=r, column=commnet_column_number).fill = yellow_fill
        self.workbook.save(self.url)


    def get_img_path(self, path):
        img_path = ""
        return img_path

    def get_all_path(self,path_total):
        # 1、获取总目录下所有的文件夹名称
        paths = os.listdir(path_total)
        for p in range(len(paths)):
            # 2、获取指定目录下，所有文件及文件夹名称
            # 3、获取csv文件的路径
            # 4、获取图片所在的路径
            path = paths[p]

            file_list = os.listdir(path)
            folder_list = [os.path.join(path, item) for item in file_list if os.path.isdir(os.path.join(path, item))]
            for f in range(len(folder_list)):
                img_path = self.get_img_path(f)
class easy_click(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Easy Click")

        # 创建第一个分区
        partition1 = ttk.Labelframe(self, text="外业采集数据转换")

        self.label1 = tk.Label(partition1, text="图片目录路径:")
        self.label1.grid(row=0, column=0, padx=10, pady=10, sticky="w")  # 放置在第0行第0列

        self.entry_var1 = tk.StringVar()
        self.entry1 = tk.Entry(partition1, textvariable=self.entry_var1, width=50)
        self.entry1.grid(row=0, column=1, padx=10, pady=10, sticky="w")  # 放置在第0行第1列

        self.button1 = tk.Button(partition1, text="①选择图片目录", command=lambda: self.open_folder_dialog(self.entry_var1))
        self.button1.grid(row=0, column=2, padx=10, pady=10, sticky="w")  # 放置在第0行第2列

        # image存放的路径
        self.label2 = tk.Label(partition1, text="文件路径:")
        self.label2.grid(row=1, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

        self.entry_var2 = tk.StringVar()
        self.entry2 = tk.Entry(partition1, textvariable=self.entry_var2, width=50)
        self.entry2.grid(row=1, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

        self.button2 = tk.Button(partition1, text="②选择文件", command=lambda: self.open_file_dialog(self.entry_var2))
        self.button2.grid(row=1, column=2, padx=10, pady=10, sticky="w")  # 放置在第1行第2列

        # 定义一个变量，用于保存选择框的状态
        self.selected_subject_option = tk.StringVar()
        self.selected_subject_option.set("2")
        self.subject_names = [('市政', "1"), ('路灯', "2"), ('保洁', "3"), ('园林', "4"), ('其他', "5")]

        for text, value in self.subject_names:
            self.subject_button = tk.Radiobutton(partition1, text=text, variable=self.selected_subject_option, value=value)
            self.subject_button.grid(row=2, column=int(value) - 1, padx=10, pady=10, sticky='w')

        # image存放的路径
        self.belong_to_label = tk.Label(partition1, text="③设施归属:")
        self.belong_to_label.grid(row=3, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

        # belong_to_var = tk.StringVar()

        # self.belong_to_combox = ttk.Combobox(partition1, values=['东山街道（区管范围）',
        #                                               '东山街道',
        #                                               '秣陵街道',
        #                                               '淳化街道',
        #                                               '麒麟街道',
        #                                               '汤山街道',
        #                                               '湖熟街道',
        #                                               '禄口街道',
        #                                               '横溪街道',
        #                                               '江宁街道',
        #                                               '谷里街道',
        #                                               '江宁开发区',
        #                                               '空港开发区',
        #                                               '江苏软件园',
        #                                               '江宁高新区',
        #                                               '滨江开发区',
        #                                               '未来科技城',
        #                                               '麒麟科创园',
        #                                               '东山总部园',
        #                                               '上坊片区',
        #                                               '南京南站',
        #                                               '牛首山',
        #                                               '园博园',
        #                                               ], width=50)
        self.belong_to_combox = ttk.Combobox(partition1, values=['区管范围',
                                                      '江宁街道',
                                                      '谷里街道',
                                                      '江宁开发区',
                                                      '空港开发区',
                                                      '江苏软件园',
                                                      '江宁高新区',
                                                      '滨江开发区',
                                                      '未来科技城',
                                                      '麒麟科创园',
                                                      '东山街道',
                                                      '东山街道（土山管办）',
                                                      '东山街道（上坊片区）',
                                                      '东山街道（南站片区）',
                                                      '牛首山',
                                                      '园博园',
                                                      '秣陵街道',
                                                      '淳化街道',
                                                      '麒麟街道',
                                                      '汤山街道',
                                                      '湖熟街道',
                                                      '禄口街道',
                                                      '横溪街道',
                                                      ], width=50)
        self.belong_to_combox.current(0)
        # belong_to_entry = tk.Entry(root, textvariable=belong_to_var, width=50)
        self.belong_to_combox.grid(row=3, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

        # 最下方按钮
        # 创建开始转换按钮和关闭程序按钮（位于同一行）
        self.start_button = tk.Button(partition1, text="④开始转换", command=self.start_conversion)
        self.start_button.grid(row=6, column=0, padx=10, pady=10, sticky="w")  # 放置在第2行第0列
        partition1.pack(side=tk.TOP, padx=10, pady=10, fill=tk.BOTH, expand=True)

        # 创建第二个分区
        partition2 = ttk.Labelframe(self, text="道路坐标转换")
        label2 = tk.Label(partition2, text="这是第二个分区的内容")
        # 分割线文件路径
        self.change_road_coordinate_file_path = tk.Label(partition2, text="I道路文件:")
        self.change_road_coordinate_file_path.grid(row=4, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

        self.change_road_coordinate_file_var = tk.StringVar()
        self.change_road_coordinate_file_entery = tk.Entry(partition2, textvariable=self.change_road_coordinate_file_var, width=50)
        self.change_road_coordinate_file_entery.grid(row=4, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

        self.change_road_file_button = tk.Button(partition2, text="I选择道路文件",
                                            command=lambda: self.open_file_dialog(self.change_road_coordinate_file_var))
        self.change_road_file_button.grid(row=4, column=2, padx=10, pady=10, sticky="w")  # 放置在第1行第2列

        # 添加列的名称
        self.coordinate_column_label = tk.Label(partition2, text="II道路坐标所在列:")
        self.coordinate_column_label.grid(row=5, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

        self.coordinate_column_label_entry_var = tk.StringVar()
        self.coordinate_column_label_entry = tk.Entry(partition2, textvariable=self.coordinate_column_label_entry_var, width=50)
        self.coordinate_column_label_entry.grid(row=5, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

        # 定义一个变量，用于保存选择框的状态
        self.coordinate_to_plant = tk.StringVar()
        self.coordinate_to_plant.set("1")
        self.coordinate_methode = [('提取原表坐标', "1"), ('坐标向平台导入', "2")]

        for text, value in self.coordinate_methode:
            self.subject_button = tk.Radiobutton(partition2, text=text, variable=self.coordinate_to_plant,
                                                 value=value)
            self.subject_button.grid(row=2, column=int(value) - 1, padx=10, pady=10, sticky='w')

        self.start_button = tk.Button(partition2, text="III道路坐标转换", command=self.change_road_coordinate)
        self.start_button.grid(row=6, column=0, padx=10, pady=10, sticky="w")  # 放置在第2行第0列

        partition2.pack(side=tk.TOP, padx=10, pady=10, fill=tk.BOTH, expand=True)

        # 创建第二个分区
        partition3 = ttk.Labelframe(self, text="统一文件名后缀")
        self.clean_path_label = tk.Label(partition3, text="文件目录路径:")
        self.clean_path_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")  # 放置在第0行第0列

        self.clean_path_var1 = tk.StringVar()
        self.clean_path = tk.Entry(partition3, textvariable=self.clean_path_var1, width=50)
        self.clean_path.grid(row=0, column=1, padx=10, pady=10, sticky="w")  # 放置在第0行第1列

        self.clean_path = tk.Button(partition3, text="①选择文件目录",
                                 command=lambda: self.open_folder_dialog(self.clean_path_var1))
        self.clean_path.grid(row=0, column=2, padx=10, pady=10, sticky="w")  # 放置在第0行第2列

        self.start_button = tk.Button(partition3, text="③统一文件后缀", command=self.clean_path_ext)
        self.start_button.grid(row=6, column=0, padx=10, pady=10, sticky="w")  # 放置在第2行第0列

        partition3.pack(side=tk.TOP, padx=10, pady=10, fill=tk.BOTH, expand=True)

        # 创建第四个分区，
        partition4 = ttk.Labelframe(self, text="文件自动转换")

        # 定义一个变量，用于保存选择框的状态
        self.put_img_in_selected_subject_option = tk.StringVar()
        self.put_img_in_selected_subject_option.set("1")
        self.put_img_in_subject_names = [('道路照片', "1"), ('桥梁照片', "2"), ('公厕照片', "3")]

        for text, value in self.put_img_in_subject_names:
            self.put_img_in_subject_button = tk.Radiobutton(partition4, text=text, variable=self.put_img_in_selected_subject_option,
                                                 value=value)
            self.put_img_in_subject_button.grid(row=1, column=int(value) - 1, padx=10, pady=10, sticky='w')

        # image存放的路径
        self.standard_road_label = tk.Label(partition4, text="总表:")
        self.standard_road_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

        self.standard_road_var = tk.StringVar()
        self.standard_road_entry = tk.Entry(partition4, textvariable=self.standard_road_var, width=50)
        self.standard_road_entry.grid(row=2, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

        self.standard_road_button = tk.Button(partition4, text="②选择总表", command=lambda: self.open_file_dialog(self.standard_road_var))
        self.standard_road_button.grid(row=2, column=2, padx=10, pady=10, sticky="w")  # 放置在第1行第2列

        # image存放的路径
        self.road_to_be_confirmed_label = tk.Label(partition4, text="导出文件:")
        self.road_to_be_confirmed_label.grid(row=3, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

        self.road_to_be_confirmed_var = tk.StringVar()
        self.road_to_be_confirmed_entry = tk.Entry(partition4, textvariable=self.road_to_be_confirmed_var, width=50)
        self.road_to_be_confirmed_entry.grid(row=3, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

        self.road_to_be_confirmed_button = tk.Button(partition4, text="②选择导出文件", command=lambda: self.open_file_dialog(self.road_to_be_confirmed_var))
        self.road_to_be_confirmed_button.grid(row=3, column=2, padx=10, pady=10, sticky="w")  # 放置在第1行第2列

        self.start_button = tk.Button(partition4, text="③生成", command=self.mkdir_put_img)
        self.start_button.grid(row=4, column=0, padx=10, pady=10, sticky="w")  # 放置在第2行第0列

        partition4.pack(side=tk.TOP, padx=10, pady=10, fill=tk.BOTH, expand=True)

        # 创建第五个分区，
        partition5 = ttk.Labelframe(self, text="道路指定照片")

        # 定义一个变量，用于保存选择框的状态
        self.change_road_img_var = tk.StringVar()
        self.change_road_img_var.set("1")
        self.change_road_img_methode = [('路名转换为road_id', "1"), ('向文件夹内添加照片', "2")]

        for text, value in self.change_road_img_methode:
            self.subject_button = tk.Radiobutton(partition5, text=text, variable=self.change_road_img_var,
                                                 value=value)
            self.subject_button.grid(row=1, column=int(value) - 1, padx=10, pady=10, sticky='w')

        # image存放的路径
        self.base_road_label = tk.Label(partition5, text="道路文件:")
        self.base_road_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

        self.base_road_var = tk.StringVar()
        self.base_road_entry = tk.Entry(partition5, textvariable=self.base_road_var, width=50)
        self.base_road_entry.grid(row=2, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

        self.base_road_button = tk.Button(partition5, text="②选择业主提供文件", command=lambda: self.open_file_dialog(self.base_road_var))
        self.base_road_button.grid(row=2, column=2, padx=10, pady=10, sticky="w")  # 放置在第1行第2列

        # image存放的路径
        self.road_to_be_confirmed_label = tk.Label(partition5, text="照片路径:")
        self.road_to_be_confirmed_label.grid(row=3, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

        self.open_picture_path_var = tk.StringVar()
        self.open_picture_path_entry = tk.Entry(partition5, textvariable=self.open_picture_path_var, width=50)
        self.open_picture_path_entry.grid(row=3, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列
        self.open_picture_path_button = tk.Button(partition5, text="②选择照片路径", command=lambda: self.open_folder_dialog(self.open_picture_path_var))
        self.open_picture_path_button.grid(row=3, column=2, padx=10, pady=10, sticky="w")  # 放置在第1行第2列

        self.start_add_button = tk.Button(partition5, text="③开始转换", command=self.change_file_name_and_get_excel)
        self.start_add_button.grid(row=4, column=0, padx=10, pady=10, sticky="w")  # 放置在第2行第0列

        partition5.pack(side=tk.TOP, padx=10, pady=10, fill=tk.BOTH, expand=True)

        # partition3 =
        self.close_button = tk.Button(self, text="关闭程序", command=self.quit)
        self.close_button.pack(side=tk.TOP, padx=10, pady=10, fill=tk.BOTH, expand=True)

    # 定义函数：打开目录选择对话框，并在输入框中显示文件夹路径
    def open_folder_dialog(self, entry_var):
        folder_path = filedialog.askdirectory()  # 打开目录选择对话框
        if folder_path:  # 如果用户选择了目录
            entry_var.set(folder_path)  # 在输入框中显示目录路径

    # 定义函数：打开文件选择对话框，并在输入框中显示文件路径
    def open_file_dialog(self,entry_var):
        file_path = filedialog.askopenfilename()  # 打开文件选择对话框
        if file_path:  # 如果用户选择了文件
            entry_var.set(file_path)  # 在输入框中显示文件路径

    # 定义函数：开始转换的操作（示例函数，需要根据实际需求编写）
    def start_conversion(self):
        rcf = ReadAndCompareFileName()
        # entry_var2.set("E:/项目文件夹/江宁普查项目外业资料/测试资料/路灯/北沿路-照明-表格.csv")
        # entry_var1.set("E:/项目文件夹/江宁普查项目外业资料/测试资料/路灯/北沿路")
        if self.entry_var2.get() is None or self.entry_var2.get() == "":
            messagebox.showinfo("woring！！！", "请选择.xls、.xlsx、.csv文件")
        elif self.entry_var1.get() is None or self.entry_var1.get() == "":
            messagebox.showinfo("woring！！！", "请选择照片所在文件夹")
        else:
            # 进行文件转换操作
            url_excel = None
            if self.entry_var2.get().endswith(".csv"):
                # 读取CSV文件
                csv_file = self.entry_var2.get()
                detect_encoding = rcf.detect_encoding(csv_file)

                if detect_encoding == 'ISO-8859-1' or detect_encoding == 'MacRoman' or detect_encoding == "GB2312":
                    detect_encoding = 'gbk'

                df = pd.read_csv(csv_file, encoding=detect_encoding)
                # 将DataFrame保存为Excel文件
                url_excel = self.entry_var2.get().replace(".csv", "") + '.xlsx'
                df.to_excel(url_excel, index=False)
            elif self.entry_var2.get().endswith(".xlsx"):
                url_excel = self.entry_var2.get()
            elif self.entry_var2.get().endswith(".xls"):
                url_excel = self.entry_var2.get()
            else:
                messagebox.showinfo("woring！！！", "无法打开所选文件，请重新选择")
            # # addName = add_column_label_entry.get()
            # if addName == "" or addName is None:
            addName = "img_path"
            data = rcf.get_data(url_excel, addName)
            # 获取目录下文件名称
            url_img = self.entry_var1.get()
            img_name = os.listdir(url_img)
            belong_to = self.belong_to_combox.get()
            if belong_to is None or belong_to == "":
                belong_to = "东山街道（区管范围）"
            rcf.add_belong_to(belong_to)
            if self.selected_subject_option.get() == "1":
                '''
                按照表格将内容进行转换
                '''
                rcf.field_matching()
                rcf.get_image_path_to_excel(data, img_name, url_img)
                rcf.add_point_image_id('point_img_id_name', 'point_img_id_value')
            elif self.selected_subject_option.get() == "2":
                '''
                提取备注中灯头数量
                '''
                rcf.add_ludeng_column()
                rcf.get_image_path_to_excel(data, img_name, url_img)
                rcf.add_point_image_id('point_img_id_name', 'point_img_id_value')
                rcf.set_back_color()
            elif self.selected_subject_option.get() == "3":
                '''
                直接进行转换即可
                '''
                rcf.clean_huanwei_column()
                rcf.get_image_path_to_excel(data, img_name, url_img)
                rcf.add_point_image_id('point_img_id_name', 'point_img_id_value')
            elif self.selected_subject_option.get() == "4":
                rcf.get_image_path_to_excel(data, img_name, url_img)
                rcf.add_point_image_id('point_img_id_name', 'point_img_id_value')
            elif self.selected_subject_option.get() == "5":
                rcf.get_image_path_to_excel(data, img_name, url_img)
                rcf.add_point_image_id('point_img_id_name', 'point_img_id_value')
            rcf.my_close_workbook("excelTimeCode")
            messagebox.showinfo("转换完成", url_excel + "转换已完成！")

    def change_road_coordinate(self):
        '''
        按照专业对道路进行分割,
        同一个断面只有一个分割线。
        :return:
        '''
        global url_excel
        if self.coordinate_to_plant.get() == "1":
            rcf = ReadAndCompareFileName()
            if self.change_road_coordinate_file_var.get() is None:
                pass
            else:
                self.change_table_ext(rcf)
                rcf.read_excel(url_excel)
            messagebox.showinfo("转换完成", url_excel + "转换已完成！")
        elif self.coordinate_to_plant.get() == "2":
            rcf = ReadAndCompareFileName()
            if self.change_road_coordinate_file_var.get() is None:
                pass
            else:
                self.change_table_ext(rcf)
                rcf.read_excel2(url_excel)
            messagebox.showinfo("转换完成", url_excel + "转换已完成！")


    def change_table_ext(self, rcf):
        '''
        修改文件后缀
        :param rcf:
        :return:
        '''
        global url_excel
        if self.change_road_coordinate_file_var.get().endswith(".csv"):
            # 读取CSV文件
            csv_file = self.change_road_coordinate_file_var.get()
            detect_encoding = rcf.detect_encoding(csv_file)
            if detect_encoding == 'ISO-8859-1' or detect_encoding == 'MacRoman':
                detect_encoding = 'gbk'
            df = pd.read_csv(csv_file, encoding=detect_encoding)
            # 将DataFrame保存为Excel文件

            path_name = os.path.dirname(csv_file)
            (file, ext) = os.path.splitext(csv_file)
            file_name = os.path.basename(csv_file).replace(ext, "") + "-提交平台"
            url_excel = path_name + "/" + file_name + ".xlsx"
            df.to_excel(url_excel, index=False)
        elif self.change_road_coordinate_file_var.get().endswith(".xlsx"):
            url_excel = self.change_road_coordinate_file_var.get()
        elif self.change_road_coordinate_file_var.get().endswith(".xls"):
            url_excel = self.change_road_coordinate_file_var.get()
        else:
            messagebox.showinfo("woring！！！", "无法打开所选文件，请重新选择")

    def get_last_road(self):
        '''
        获取还剩下多少条路未在地图中
        :return:
        '''

    def clean_path_ext(self):
        '''
        统一文件后缀
        :return:
        '''
        file_path = self.clean_path_var1.get()
        file_list = os.listdir(file_path)
        for i in range(len(file_list)):
            file_full_path = os.path.join(file_path, file_list[i])
            (file_name,ext) = os.path.splitext(file_full_path)
            if ext != ".ovobj":
                os.rename(file_full_path,file_name + ".ovobj")
        messagebox.showinfo("转换完成", file_path + "转换已完成！")

    def mkdir_put_img(self):
        '''
        根据选择，使用不同的方法
        :return:
        '''

        select_options = self.put_img_in_selected_subject_option.get()
        # 总表
        total_excel = self.standard_road_entry.get()
        total_workbook = openpyxl.load_workbook(total_excel)

        # 导出数据
        out_excel = self.road_to_be_confirmed_var.get()
        out_pd = pd.read_excel(out_excel)

        if select_options == "1":
            # 获取道路表格
            total_sheetname = total_workbook.sheetnames[0]
            total_sheet = total_workbook[total_sheetname]
            max_row = total_sheet.max_row
            farther_path = os.path.dirname(out_excel)
            os.mkdir(farther_path + "/道路照片")
            for r in range(5, max_row+1):
                road_id = total_sheet.cell(row = r, column = 1).value

                road_name = total_sheet.cell(row = r, column = 4).value
                out_data = out_pd[out_pd["道路名称"] == road_name]

                # 创建文件夹

                put_img_path= farther_path + "/道路照片/" + road_id

                os.mkdir(put_img_path)
                if out_data.index.shape[0] > 0:
                    ink_fill = PatternFill(start_color="FF0066", end_color="FF0066", fill_type="solid")
                    total_sheet.cell(row=r, column=3).fill = ink_fill
                    # 获取所有对应道路的照片路径
                    for index in out_data.index:
                        img_data = farther_path +"/"+ out_data["img_path"][index].replace("./","/")
                        (path, filename) = os.path.split(img_data)
                        shutil.copyfile(img_data, put_img_path + "/" + filename)
        elif select_options == "2":
            sheetname = total_workbook.sheetnames[0]
            # 获取道路表格
            total_sheetname = total_workbook.sheetnames[0]
            total_sheet = total_workbook[total_sheetname]
            max_row = total_sheet.max_row
            for r in range(5, max_row):
                road_id = total_sheet.cell(row=r, column=1).value

                road_name = total_sheet.cell(row=r, column=4).value
                out_data = out_pd[out_pd["道路名称"] == road_name]

                # 创建文件夹
                farther_path = os.path.dirname(out_excel)
                put_img_path = farther_path + "/桥梁照片/" + road_id
                os.mkdir(farther_path + "/桥梁照片")
                os.mkdir(put_img_path)

                # 获取所有对应道路的照片路径
                for index in out_data:
                    img_data = farther_path + out_data[index]["img_path"].replace("./", "/")
                    shutil.copyfile(img_data, put_img_path)
            print("桥梁")
        elif select_options == "3":
            sheetname = total_workbook.sheetnames[0]
            # 获取道路表格
            total_sheetname = total_workbook.sheetnames[0]
            total_sheet = total_workbook[total_sheetname]
            max_row = total_sheet.max_row
            for r in range(5, max_row):
                road_id = total_sheet.cell(row=r, column=1).value

                road_name = total_sheet.cell(row=r, column=4).value
                out_data = out_pd[out_pd["道路名称"] == road_name]

                # 创建文件夹
                farther_path = os.path.dirname(out_excel)
                put_img_path = farther_path + "/公厕照片/" + road_id
                os.mkdir(farther_path + "/公厕照片")
                os.mkdir(put_img_path)

                # 获取所有对应道路的照片路径
                for index in out_data:
                    img_data = farther_path + out_data[index]["img_path"].replace("./", "/")
                    shutil.copyfile(img_data, put_img_path)
        messagebox.showinfo("转换已完成！")

    def compare_standard_confirmed(self):
        '''
        比较两个表格间存在的问题
        :return:
        '''
        standard_road = self.standard_road_entry.get()
        road_confirmed = self.road_to_be_confirmed_var.get()

        standard_road = pd.read_excel(standard_road) # 获取标准文件中的内容

        road_confirmed = pd.read_excel(road_confirmed) # 获取对比文件中的内容
        road_confirmed_list = list(road_confirmed.groupby(by="文件名称").indices.keys())
        road_confirmed_df = pd.DataFrame({"roadname":road_confirmed_list})
        not_in_list = list()
        save_to_excel = list()
        for i in range(len(standard_road["道路名称"])):
            for j in range(len(road_confirmed_list)):
                if standard_road["道路名称"][i] not in road_confirmed_list[j]:
                    not_in_list.append(standard_road["道路名称"][i])
            if len(not_in_list) == len(road_confirmed_list):
                save_to_excel.append(standard_road["道路名称"][i])
        save_to_excel_dict = dict()
        save_to_excel_dict["剩余道路"] = save_to_excel
        result_df = pd.DataFrame(save_to_excel_dict)

        result_df.to_excel("")

    def change_file_name_and_get_excel(self):
        if self.change_road_img_var.get() == "1":
            base_excel_path = self.base_road_var.get()
            picture_path = self.open_picture_path_var.get()
            road_name_list = os.listdir(picture_path)
            workbook = openpyxl.load_workbook(base_excel_path)
            # 获取sheet
            sheet_name = workbook.sheetnames[0]
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            for i in range(len(road_name_list)):
                road_name = road_name_list[i]
                for j in range(5, max_row+1):
                    if str(sheet.cell(row = j, column = 4).value).strip(" ") == road_name.split("（")[0] or str(sheet.cell(row = j, column = 4).value).strip(" ") == road_name.split("(")[0]:
                        os.rename(picture_path + "/" + road_name,picture_path + "/" + sheet.cell(row = j, column = 1).value)
                        # yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        pink_fill = PatternFill(start_color="FF0066", end_color="FF0066", fill_type="solid")
                        sheet.cell(row=j, column=1).fill = pink_fill
                        workbook.save(base_excel_path)
                    if "（" in road_name or "(" in road_name:
                        pink_fill = PatternFill(start_color="FF0066", end_color="FF0066", fill_type="solid")
                        sheet.cell(row=j, column=1).fill = pink_fill
                        break
            workbook.save(base_excel_path)
            # 获取id并且，将其转换为excel表格
            self.add_picture_to_excel()
        else:
            base_excel_path = self.base_road_var.get() # 市政交通表格
            farther_path = os.path.abspath(os.path.dirname(base_excel_path) + os.path.sep + ".")
            img_excel_data = pd.read_excel(base_excel_path)
            put_img_to_file_path = self.open_picture_path_var.get() # 需要修改的所有文件的路径
            file_and_dir = os.listdir(put_img_to_file_path)
            excel_file_path = str()

            self.put_zhilubiaozhi_img(farther_path, img_excel_data, put_img_to_file_path)
            messagebox.showinfo("转换已完成！")
            pass
        pass

    def put_zhilubiaozhi_img(self, farther_path, img_excel_data, put_img_to_file_path):
        '''
        遍历表格，并将指路标志复制到指定目录下
        :param farther_path:
        :param img_excel_data:
        :param put_img_to_file_path:
        :return:
        '''
        file_and_dir = os.listdir(put_img_to_file_path)
        img_dir_list = list()
        excel_file_path = str()
        for i in file_and_dir:
            name = put_img_to_file_path + "/" + i
            if os.path.isfile(name):
                excel_file_path = name
            else:
                img_dir_list.append(i)
        # 获取主表内数据
        work_book = openpyxl.load_workbook(excel_file_path)
        sheet_name = work_book.sheetnames[0]
        sheet = work_book[sheet_name]
        max_row = sheet.max_row
        # 遍历dir，获取道路名称
        for i in range(len(img_dir_list)):
            road_id = img_dir_list[i]
            # 遍历表格
            for r in range(5, max_row + 1):
                if sheet.cell(row=r, column=1).value == road_id:
                    road_name = sheet.cell(row=r, column=4).value
                    data_series = img_excel_data[
                        (img_excel_data["道路名称"] == road_name) & ((img_excel_data["Comment"] == "指路标志") | (img_excel_data["Comment"] == "指路标志"))]
                    if data_series.shape[0] != 0:
                        pink_fill = PatternFill(start_color="FF0066", end_color="FF0066", fill_type="solid")
                        sheet.cell(row=r, column=2).fill = pink_fill
                        work_book.save(excel_file_path)
                        for img_relpath_index in data_series["img_path"]:
                            old_path = farther_path + "/" + img_relpath_index.replace("./", "")
                            (path, filename) = os.path.split(old_path)
                            new_path = put_img_to_file_path + "/" + road_id + "/" + filename
                            shutil.copyfile(old_path, new_path)
                pass

    def add_picture_to_excel(self):
        '''
        向道路文件夹内添加照片
        :return:
        '''
        picture_path = self.open_picture_path_var.get()
        road_id_list = os.listdir(picture_path)
        to_excel = dict()
        road_ids = list()
        img_paths = list()
        for i in range(len(road_id_list) ):
            road_id = road_id_list[i]
            road_id_base = road_id_list[i]
            img_path = ""
            img_path = picture_path + "/" + road_id
            if os.path.isfile(img_path):
                continue
            img_path_list = os.listdir(img_path)
            if "（" in road_id:
                road_id = road_id.split("（")[0]
            elif "(" in road_id:
                road_id = road_id.split("(")[0]
            elif "-" in road_id:
                road_id = road_id.split("_")[0]
            for j in range(len(img_path_list)):
                # grader_father = os.path.abspath(os.path.dirname(img_path + "/" + img_path_list[j]) + os.path.sep + "..").replace("\\","/")
                # img_path = img_path + "/" + img_path_list[j]
                img_paths.append( "./" +road_id_base+"/"+ img_path_list[j] )
                road_ids.append(road_id)
        to_excel["road_id"] = road_ids
        to_excel["img_path"] = img_paths
        df = pd.DataFrame(to_excel)
        df.to_excel(picture_path + "/" + "表格.xlsx",index=False)
        messagebox.showinfo("转换完成", picture_path + "/" + "表格.xlsx" + "转换已完成！")

if __name__ == "__main__":
    app = easy_click()
    app.mainloop()