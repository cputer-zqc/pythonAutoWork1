import openpyxl
import pandas as pd
import os


class ReadAndCompareFileName:

    def __init__(self):
        self.workbook = None
        self.sheet_name = None
        self.url = None
        self.picture_number = dict()

    def __get_max_column_row(self):
        '''
        获取当前表格醉倒列与最大行数
        :return: {"mr":max_row, "mc":max_column}
        '''
        sheet = self.workbook[self.sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        return [max_row, max_column]

    def __get_excel_data_code(self, data, i):
        '''
        获取excel表格中时间字符
        :param data:
        :param i:
        :return:
        '''
        print("正在修改第" + data["名称"][i] + "对象的图片，请稍后... ...")
        name_split = data["名称"][i].split("_")
        times = ""
        if (":" in name_split[len(name_split) - 1]):
            # 获取名称中的时间
            times = name_split[len(name_split) - 1].split(":")
        else:
            times = name_split[len(name_split) - 1]
        time = ""
        for t in times:
            time = time + t
        if len(time) > 15:
            time = time[0: 16]
        excel_img_time_code = name_split[1] + "_" + time
        return excel_img_time_code

    def get_data(self, url):
        '''
        获取表格中信息
        :param url:
        :return:
        '''
        self.url = url
        # 获取表格
        self.workbook = openpyxl.load_workbook(self.url)

        # 获取sheet
        self.sheet_name = self.workbook.sheetnames[0]
        sheet = self.workbook[self.sheet_name]
        if "excelTimeCode" not in sheet[1]:
            [max_row, max_column] = self.__get_max_column_row()
            # 添加空列
            for row_index in range(1, max_row + 1):
                if row_index == 1:
                    sheet.cell(row=row_index, column=max_column + 1, value="excelTimeCode")
                else:
                    sheet.cell(row=row_index, column=max_column + 1, value=None)
        self.workbook.save(url)
        # 数据转换
        data = sheet.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        # 返回数据
        return df

    def compare_change(self, data, img_name, beforString, imgURL):
        '''
        比较并修改文件名
        :param data:
        :param img_name:
        :param beforString:
        :param imgURL:
        :return:
        '''
        imgURL = imgURL + "/"
        # 遍历并匹配
        for i in range(0, len(data["名称"])):
            excel_img_time_code = self.__get_excel_data_code(data, i)
            # 照片编号
            number = 1
            # 遍历照片名称
            for j in range(0, len(img_name)):
                # 只考虑一个对象上绑定了多个照片。
                name = img_name[j].replace(".png", "")
                img_time_code = ""
                # 名称与照片名称中的时间进行匹配
                img_split = name.split("_")
                if ("(" in (img_split[1] + "_" + img_split[len(img_split) - 1])):
                    img_time_code = (img_split[1] + "_" + img_split[2]).split("(")[0]
                else:
                    img_time_code = (img_split[1] + "_" + img_split[2])
                img_time_code = img_time_code[0: 15]
                imgname = ""
                if excel_img_time_code == img_time_code:
                    name = name + ".png"
                    imgname = beforString + img_time_code + "_lon_" + str(data["经度"][i]) + "_lat_" + str(
                        data["纬度"][i])
                    # 表格中每一行只存储一个照片名称
                    if number <= 1:
                        [max_row, max_column] = self.__get_max_column_row()
                        sheet = self.workbook[self.sheet_name]
                        # for r in range(2, max_row + 1):
                        sheet.cell(row=i+2, column=max_column, value=imgname)
                        self.workbook.save(self.url)
                    self.picture_number[imgname] = number
                    os.rename(imgURL + name, imgURL + imgname + "-" + str(number) + ".png")
                    number += 1

    def updata_img_path_excel(self, img_url, column_name):
        # 列不存在时，进行添加，当列存在时，不进行添加
        sheet = self.workbook[self.sheet_name]
        [max_row, max_column] = self.__get_max_column_row()
        if "图片路径" not in sheet[1]:
            for r in range(1, max_row):
                if (r == 1):
                    sheet.cell(row = r, column = max_column, value = "column_name")
                else :
                    sheet.cell(row = r, column = max_column, value = None)
        # 向其中添加数据

    def save_excel(self,img_url):
        print("开始修改表格")
        sheet = self.workbook[self.sheet_name]
        [max_row, max_column] = self.__get_max_column_row()
        # 迭代函数每保存一次就需要重新遍历一次但是要记录上次添加数据的最终位置
        img_name_list = os.listdir(img_url)
        has_add_name = []
        r = 2
        while r < sheet.max_column:
            addNumber = self.picture_number[sheet.cell(row = r, column = max_column).value]
            cNumber = 0
            # 插入与上一行相同的数据
            if addNumber > 1:
                # 获取上一行数据
                previous_rows = list(sheet.iter_rows(min_row=r, max_row=r, values_only= True))[0]
                # 在指定行插入与上一行相同的数据
                sheet.insert_rows(r+1,amount = addNumber-1)
                for r_insert in range(r+1, r+addNumber):
                    for c_insert in range(1, sheet.max_column+1):
                        sheet.cell(row=r_insert, column=c_insert, value=previous_rows[c_insert-1])
                        self.workbook.save(self.url)
            r = r + addNumber
        for r in range(1, sheet.max_column):
            for name in img_name_list:
                if name not in has_add_name and name.startswith(sheet.cell(row=r, column=sheet.max_column).value):
                    print("修改第"+str(r)+"行的路径\n")
                    file_path = img_url +"/"+ name
                    sheet.cell(row = r, column =sheet.max_column, value = os.path.relpath(file_path, self.url))
                    has_add_name.append(name)
                    self.workbook.save(self.url)
                    break

    def __get_column_row_number(self, value):
        '''
        查找具有指定元素的单元格
        :param value:
        :return:
        '''
        sheet = self.workbook[self.sheet_name]
        # 指定要查找的元素
        target_element = value
        # 遍历工作表，查找具有指定元素的单元格
        for row_index, row in enumerate(sheet.iter_rows(), start=1):
            for col_index, cell in enumerate(row, start=1):
                if cell.value == target_element:
                    # 找到了具有指定元素的单元格
                    return [row_index, col_index ]

    def my_close_workbook(self, value):
        sheet = self.workbook[self.sheet_name]
        if value in sheet[1]:
            [index_row, index_column] = self.__get_column_row_number(value)
            sheet.delete_cols(index_column)
        self.workbook.close()
# if __name__ == "__main__":
#     # 获取Excel表格中数据
#     url_excel = '../doc/我的照片与录音.xlsx'
#     data = getData(url_excel)
#
#     # 获取目录下文件名称
#     url_img = '../doc/我的照片与录音'
#     img_name = os.listdir(url_img)
#
#     beforString = "照片_"
#     compareAndChange(data, img_name, beforString)
