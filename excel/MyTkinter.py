import tkinter as tk
from tkinter import filedialog
import os
from tkinter import messagebox
import pandas as pd
import openpyxl
import chardet
from openpyxl.cell import _writer


class ReadAndCompareFileName:

    def __init__(self):
        self.workbook = None
        self.sheet_name = None
        self.url = None
        self.picture_number = dict()

    def __get_max_column_row(self):
        '''
        获取当前表格醉倒列与最大行数
        :return: {"mr":max_row, "mc":max_column}
        '''
        sheet = self.workbook[self.sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        return [max_row, max_column]

    def __get_excel_data_code(self, data, i):
        '''
        获取excel表格中时间字符
        :param data:
        :param i:
        :return:
        '''
        print("正在修改第" + data["名称"][i] + "对象的图片，请稍后... ...")
        name_split = data["名称"][i].split("_")
        times = ""
        if (":" in name_split[len(name_split) - 1]):
            # 获取名称中的时间
            times = name_split[len(name_split) - 1].split(":")
        else:
            times = name_split[len(name_split) - 1]
        time = ""
        for t in times:
            time = time + t
        if len(time) > 15:
            time = time[0: 16]
        excel_img_time_code = name_split[1] + "_" + time
        return excel_img_time_code

    def detect_encoding(self, file_path):
        with open(file_path, 'rb') as f:
            detector = chardet.universaldetector.UniversalDetector()
            for line in f.readlines():
                detector.feed(line)
                if detector.done:
                    break
            detector.close()
            encoding = detector.result['encoding']
        return encoding

    def get_data(self, url, addName):
        '''
        获取表格中信息
        :param url:
        :return:
        '''
        self.url = url
        # 获取表格
        self.workbook = openpyxl.load_workbook(self.url)

        # 获取sheet
        self.sheet_name = self.workbook.sheetnames[0]
        sheet = self.workbook[self.sheet_name]
        row_data = []
        for cell in sheet[1]:
            row_data.append(cell.value)
        if addName not in row_data[1]:
            [max_row, max_column] = self.__get_max_column_row()
            # 添加空列
            for row_index in range(1, max_row + 1):
                if row_index == 1:
                    sheet.cell(row=row_index, column=max_column + 1, value=addName)
                else:
                    sheet.cell(row=row_index, column=max_column + 1, value=None)
        self.workbook.save(url)
        # 数据转换
        data = sheet.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        # 返回数据
        return df

    def compare_change(self, data, img_name, beforString, imgURL):
        '''
        比较并修改文件名
        :param data:
        :param img_name:
        :param beforString:
        :param imgURL:
        :return:
        '''
        imgURL = imgURL + "/"
        # 遍历并匹配
        for i in range(0, len(data["名称"])):
            excel_img_time_code = self.__get_excel_data_code(data, i)
            # 照片编号
            number = 1
            # 遍历照片名称
            for j in range(0, len(img_name)):
                name, file_extension = os.path.splitext(img_name[j])
                # 只考虑一个对象上绑定了多个照片。
                # name = img_name[j].replace(".png", "")
                img_time_code = ""
                # 名称与照片名称中的时间进行匹配
                img_split = name.split("_")
                if ("(" in (img_split[1] + "_" + img_split[len(img_split) - 1])):
                    img_time_code = (img_split[1] + "_" + img_split[2]).split("(")[0]
                else:
                    img_time_code = (img_split[1] + "_" + img_split[2])
                img_time_code = img_time_code[0: 15]
                imgname = ""
                if excel_img_time_code == img_time_code:
                    name = name + file_extension
                    imgname = beforString + img_time_code + "_lon_" + str(data["经度"][i]) + "_lat_" + str(
                        data["纬度"][i])
                    # 表格中每一行只存储一个照片名称
                    if number <= 1:
                        [max_row, max_column] = self.__get_max_column_row()
                        sheet = self.workbook[self.sheet_name]
                        # for r in range(2, max_row + 1):
                        sheet.cell(row=i + 2, column=max_column, value=imgname)
                        self.workbook.save(self.url)
                    self.picture_number[imgname] = number
                    os.rename(imgURL + name, imgURL + imgname + "-" + str(number) + file_extension)
                    number += 1

    def updata_img_path_excel(self, img_url, column_name):
        '''
        没用上
        :param img_url:
        :param column_name:
        :return:
        '''
        # 列不存在时，进行添加，当列存在时，不进行添加
        sheet = self.workbook[self.sheet_name]
        [max_row, max_column] = self.__get_max_column_row()
        if "图片路径" not in sheet[1]:
            for r in range(1, max_row):
                if (r == 1):
                    sheet.cell(row=r, column=max_column, value="column_name")
                else:
                    sheet.cell(row=r, column=max_column, value=None)
        # 向其中添加数据

    def save_excel(self, img_url):
        '''
        图片与excel表格中信息进行匹配，并存入图片名称
        :param img_url:
        :return:
        '''
        print("开始修改表格")
        sheet = self.workbook[self.sheet_name]
        [max_row, max_column] = self.__get_max_column_row()
        # 迭代函数每保存一次就需要重新遍历一次但是要记录上次添加数据的最终位置
        img_name_list = os.listdir(img_url)
        has_add_name = []
        r = 1
        while r <= sheet.max_row:
            addNumber = self.picture_number[sheet.cell(row=r, column=max_column).value]
            # cNumber = 0
            # 插入与上一行相同的数据
            if addNumber > 1:
                # 获取上一行数据
                previous_rows = list(sheet.iter_rows(min_row=r, max_row=r, values_only=True))[0]
                # 在指定行插入与上一行相同的数据

                sheet.insert_rows(r + 1, amount=addNumber - 1)
                for r_insert in range(r + 1, r + addNumber):
                    for c_insert in range(1, sheet.max_column + 1):
                        sheet.cell(row=r_insert, column=c_insert, value=previous_rows[c_insert - 1])
                        self.workbook.save(self.url)
                        sheet = self.workbook[self.sheet_name]
            r = r + addNumber
        for r in range(2, sheet.max_row + 1):
            for name in img_name_list:
                if name not in has_add_name and name.startswith(sheet.cell(row=r, column=sheet.max_column).value):
                    print("修改第" + str(r) + "行的路径\n")
                    file_path = img_url + "/" + name
                    sheet.cell(row=r, column=sheet.max_column, value=os.path.relpath(file_path, self.url))
                    has_add_name.append(name)
                    self.workbook.save(self.url)
                    break

    def __get_column_row_number(self, value):
        '''
        查找具有指定元素的单元格
        :param value:
        :return:
        '''
        sheet = self.workbook[self.sheet_name]
        # 指定要查找的元素
        target_element = value
        # 遍历工作表，查找具有指定元素的单元格
        for row_index, row in enumerate(sheet.iter_rows(), start=1):
            for col_index, cell in enumerate(row, start=1):
                if cell.value == target_element:
                    # 找到了具有指定元素的单元格
                    return [row_index, col_index]

    def my_close_workbook(self, value):
        # sheet = self.workbook[self.sheet_name]
        # if value in sheet[1]:
        #     [index_row, index_column] = self.__get_column_row_number(value)
        #     sheet.delete_cols(index_column)
        self.workbook.close()


# 创建主窗口
root = tk.Tk()
root.title("文件加载与转换程序")


# 定义函数：打开目录选择对话框，并在输入框中显示文件夹路径
def open_folder_dialog(entry_var):
    folder_path = filedialog.askdirectory()  # 打开目录选择对话框
    if folder_path:  # 如果用户选择了目录
        entry_var.set(folder_path)  # 在输入框中显示目录路径


# 定义函数：打开文件选择对话框，并在输入框中显示文件路径
def open_file_dialog(entry_var):
    file_path = filedialog.askopenfilename()  # 打开文件选择对话框
    if file_path:  # 如果用户选择了文件
        entry_var.set(file_path)  # 在输入框中显示文件路径


# 定义函数：开始转换的操作（示例函数，需要根据实际需求编写）
def start_conversion():
    rcf = ReadAndCompareFileName()
    if entry_var2.get() == None or entry_var2.get() == "":
        messagebox.showinfo("woring！！！", "请选择.xls、.xlsx、.csv文件")
    elif entry_var1.get() == None or entry_var1.get() == "":
        messagebox.showinfo("woring！！！", "请选择照片所在文件夹")
    else:
        # 进行文件转换操作
        url_excel = None
        if entry_var2.get().endswith(".csv"):
            # 读取CSV文件
            csv_file = entry_var2.get()
            detect_encoding = rcf.detect_encoding(csv_file)
            df = pd.read_csv(csv_file, encoding=detect_encoding)
            # 将DataFrame保存为Excel文件
            url_excel = entry_var2.get().replace(".csv", "") + '.xlsx'
            df.to_excel(url_excel, index=False)
        elif entry_var2.get().endswith(".xlsx"):
            url_excel = entry_var2.get()
        elif entry_var2.get().endswith(".xls"):
            url_excel = entry_var2.get()
        else:
            messagebox.showinfo("woring！！！", "无法打开所选文件，请重新选择")
        addName = add_column_label_entry.get()
        if addName == "" or addName is None:
            addName = "img_path"
        data = rcf.get_data(url_excel, addName)
        # 获取目录下文件名称
        url_img = entry_var1.get()
        img_name = os.listdir(url_img)

        beforString = entry_var3.get()
        if beforString == "" or beforString is None:
            beforString = "pic_"
        else:
            beforString += "_"
        rcf.compare_change(data, img_name, beforString, url_img)
        rcf.save_excel(url_img)
        rcf.my_close_workbook("excelTimeCode")

        messagebox.showinfo("转换完成", url_excel + "转换已完成！")

    # if url_excel:
    #     # 读取Excel文件
    #     df = pd.read_excel(url_excel)
    #     # 创建TreeView来显示数据
    #     tree = ttk.Treeview(root, columns=tuple(df.columns), show="headings", height=10)
    #     for col in df.columns:
    #         tree.heading(col, text=col)
    #         tree.column(col, width=100, anchor='center')  # 设置列宽和对齐方式
    #     for index, row in df.iterrows():
    #         tree.insert("", index, values=tuple(row))
    #     tree.grid(row=0, column=0, sticky="nsew")  # 放置Treeview
    #
    #     # 添加垂直滚动条
    #     y_scrollbar = ttk.Scrollbar(root, orient="vertical", command=tree.yview)
    #     y_scrollbar.grid(row=0, column=1, sticky="ns")  # 放置垂直滚动条
    #     tree.config(yscrollcommand=y_scrollbar.set)  # 关联Treeview和垂直滚动条
    #
    #     # 添加水平滚动条
    #     x_scrollbar = ttk.Scrollbar(root, orient="horizontal", command=tree.xview)
    #     x_scrollbar.grid(row=1, column=0, sticky="ew")  # 放置水平滚动条
    #     tree.config(xscrollcommand=x_scrollbar.set)  # 关联Treeview和水平滚动条
    #
    #     # 配置行和列的权重，使得滚动条可以随着窗口的大小自动调整
    #     root.grid_rowconfigure(0, weight=1)
    #     root.grid_columnconfigure(0, weight=1)


# Excel文件路径
# 创建标签、输入框和按钮（图片目录）
label1 = tk.Label(root, text="图片目录路径:")
label1.grid(row=0, column=0, padx=10, pady=10, sticky="w")  # 放置在第0行第0列

entry_var1 = tk.StringVar()
entry1 = tk.Entry(root, textvariable=entry_var1, width=50)
entry1.grid(row=0, column=1, padx=10, pady=10, sticky="w")  # 放置在第0行第1列

button1 = tk.Button(root, text="①选择图片目录", command=lambda: open_folder_dialog(entry_var1))
button1.grid(row=0, column=2, padx=10, pady=10, sticky="w")  # 放置在第0行第2列

# image存放的路径
label2 = tk.Label(root, text="文件路径:")
label2.grid(row=1, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

entry_var2 = tk.StringVar()
entry2 = tk.Entry(root, textvariable=entry_var2, width=50)
entry2.grid(row=1, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

button2 = tk.Button(root, text="②选择文件", command=lambda: open_file_dialog(entry_var2))
button2.grid(row=1, column=2, padx=10, pady=10, sticky="w")  # 放置在第1行第2列

# image存放的路径
label3 = tk.Label(root, text="图片名前缀:")
label3.grid(row=2, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

entry_var3 = tk.StringVar()
entry3 = tk.Entry(root, textvariable=entry_var3, width=50)
entry3.grid(row=2, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

# 添加列的名称
add_column_label = tk.Label(root, text="表格中图片所在列名称:")
add_column_label.grid(row=3, column=0, padx=10, pady=10, sticky="w")  # 放置在第1行第0列

add_column_label_entry_var = tk.StringVar()
add_column_label_entry = tk.Entry(root, textvariable=add_column_label_entry_var, width=50)
add_column_label_entry.grid(row=3, column=1, padx=10, pady=10, sticky="w")  # 放置在第1行第1列

# 最下方按钮
# 创建开始转换按钮和关闭程序按钮（位于同一行）
start_button = tk.Button(root, text="③开始转换", command=start_conversion)
start_button.grid(row=4, column=0, padx=10, pady=10, sticky="w")  # 放置在第2行第0列

close_button = tk.Button(root, text="关闭程序", command=root.quit)
close_button.grid(row=4, column=1, padx=10, pady=10, sticky="w")  # 放置在第2行第1列
# # 加载Excel表格按钮
# load_excel_button = tk.Button(root, text="Load Excel", command=load_excel)
# load_excel_button.pack(side=tk.LEFT)
# 运行主循环
root.mainloop()
